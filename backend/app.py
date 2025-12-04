from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import csv
import re
import openpyxl
from io import BytesIO

app = Flask(__name__)
CORS(app)

def load_courses():
    courses = []
    try:
        with open('output.csv', 'r', encoding='utf-8-sig') as file:
            csv_reader = csv.DictReader(file)
            for row in csv_reader:
                time_parts = row.get('時間', '/').split('/')
                courses.append({
                    'code': row.get('流水號', ''),
                    'name': row.get('名稱與備註', ''),
                    'teacher': row.get('教授', ''),
                    'credits': int(row.get('學分', '0')) if row.get('學分', '0').isdigit() else 0,
                    'time': time_parts[0] if len(time_parts) > 0 else '',
                    'classroom': time_parts[1] if len(time_parts) > 1 else row.get('教室', ''),
                    'type': row.get('選修別', ''),
                    'semester': row.get('全與半', ''),
                    'note': '',
                    'restrictions': row.get('分發條件內容', '')
                })
    except Exception as e:
        print(f"Error loading courses: {e}")

    return courses

ALL_COURSES = load_courses()
print(f"Loaded {len(ALL_COURSES)} courses")

def parse_course_time(time_string):
    weekday_map = {'一': '1', '二': '2', '三': '3', '四': '4', '五': '5'}
    slots = []
    parts = time_string.split(',')

    for part in parts:
        part = part.strip()
        match = re.match(r'([一二三四五])([0-9NABCD]+)', part)
        if match:
            day_chinese = match.group(1)
            periods = match.group(2)
            day = weekday_map.get(day_chinese)
            if day:
                for period in periods:
                    slots.append(f"{day}-{period}")

    return slots

def has_time_conflict(course_time, busy_times):
    course_slots = parse_course_time(course_time)
    return any(slot in busy_times for slot in course_slots)

def check_department_restriction(restrictions, user_department):
    if not user_department or not restrictions:
        return True

    restrictions_str = str(restrictions).lower()
    user_dept_lower = user_department.lower()

    if '系所:限' in restrictions_str:
        allowed_depts = re.findall(r'系所:限([^。]+)', restrictions_str)
        if allowed_depts:
            for dept_list in allowed_depts:
                if user_dept_lower in dept_list.lower():
                    return True
            return False

    return True

def check_grade_restriction(restrictions, user_grade):
    if not restrictions:
        return True

    grade_map = {1: '一', 2: '二', 3: '三', 4: '四', 5: '碩一', 6: '碩二'}

    restrictions_str = str(restrictions)
    user_grade_chinese = grade_map.get(user_grade, '')

    if '年級:限' in restrictions_str:
        if f'年級:限{user_grade_chinese}年級' in restrictions_str:
            return True
        grade_restrictions = re.findall(r'年級:限([^。]+)', restrictions_str)
        if grade_restrictions and user_grade_chinese:
            for grade_list in grade_restrictions:
                if user_grade_chinese in grade_list:
                    return True
            return False

    return True

@app.route('/api/filter-courses', methods=['POST'])
def filter_courses():
    try:
        data = request.get_json()
        department = data.get('department', '')
        grade = data.get('grade', 1)
        busy_times = set(data.get('busyTimes', []))

        print(f"Filter request - dept: {department}, grade: {grade}, busy times: {len(busy_times)}")

        filtered_courses = []
        for course in ALL_COURSES:
            if has_time_conflict(course['time'], busy_times):
                continue

            if department and not check_department_restriction(course.get('restrictions', ''), department):
                continue

            if not check_grade_restriction(course.get('restrictions', ''), grade):
                continue

            filtered_courses.append(course)

        print(f"Filter result: {len(filtered_courses)} courses")

        return jsonify({
            'success': True,
            'count': len(filtered_courses),
            'courses': filtered_courses
        })

    except Exception as e:
        print(f"Error filtering courses: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/courses', methods=['GET'])
def get_all_courses():
    return jsonify({
        'success': True,
        'count': len(ALL_COURSES),
        'courses': ALL_COURSES[:50]
    })

@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({
        'status': 'ok',
        'courses_loaded': len(ALL_COURSES)
    })

@app.route('/api/download-courses', methods=['POST'])
def download_courses():
    try:
        data = request.get_json()
        course_codes = data.get('courseCodes', [])

        if not course_codes:
            return jsonify({
                'success': False,
                'error': 'No course codes provided'
            }), 400

        print(f"Download request for {len(course_codes)} courses")

        wb = openpyxl.load_workbook('bigcontent.xlsx')
        ws = wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
        code_col_idx = None
        for idx, cell in enumerate(header_row, 1):
            if cell.value and '流水號' in str(cell.value):
                code_col_idx = idx
                break

        if code_col_idx is None:
            return jsonify({
                'success': False,
                'error': 'Course code column not found'
            }), 500

        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = 'Selected Courses'

        for cell in header_row:
            new_ws.cell(row=1, column=cell.column, value=cell.value)

        matched_count = 0
        for row in ws.iter_rows(min_row=2, values_only=False):
            course_code = row[code_col_idx - 1].value
            if course_code in course_codes:
                matched_count += 1
                new_row = []
                for cell in row:
                    new_ws.cell(row=matched_count + 1, column=cell.column, value=cell.value)

        output = BytesIO()
        new_wb.save(output)
        output.seek(0)

        print(f"Generated Excel with {matched_count} courses")

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='selected_courses.xlsx'
        )

    except Exception as e:
        print(f"Error downloading courses: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

if __name__ == '__main__':
    print("=" * 50)
    print("NCU Course Filter Backend Server")
    print("=" * 50)
    print(f"Loaded {len(ALL_COURSES)} courses")
    print("Server running at: http://localhost:3000")
    print("API endpoints:")
    print("  - POST /api/filter-courses")
    print("  - GET  /api/courses")
    print("  - GET  /api/health")
    print("=" * 50)

    app.run(host='0.0.0.0', port=3000, debug=True)